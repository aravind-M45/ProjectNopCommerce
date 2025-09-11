import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    """Return total number of rows in the given sheet."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    """Return total number of columns in the given sheet."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNum, colNum):
    """Read data from cell (rowNum, colNum)."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def writeData(file, sheetName, rowNum, colNum, data):
    """Write data into cell (rowNum, colNum)."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(file)


def appendRow(file, sheetName, dataList):
    """Append a new row at the end of the sheet."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.append(dataList)
    workbook.save(file)


def clearSheet(file, sheetName):
    """Clear all rows except headers in a sheet."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    workbook.save(file)


def fillGreen(file, sheetName, rowNum, colNum):
    """Fill cell with Green color (Pass)."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color="60b212", end_color="60b212", fill_type="solid")
    sheet.cell(row=rowNum, column=colNum).fill = greenFill
    workbook.save(file)


def fillRed(file, sheetName, rowNum, colNum):
    """Fill cell with Red color (Fail)."""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
    sheet.cell(row=rowNum, column=colNum).fill = redFill
    workbook.save(file)
